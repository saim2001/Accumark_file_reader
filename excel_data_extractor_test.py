import re
import json
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import pandas as pd
import os


# def read_file(file_path):
#     try:
#         # Parse the XML file
#         tree = ET.parse(file_path)
#         root = tree.getroot()

#         # Find the document_content element
#         document_content = root.find('.//document_content')

#         if document_content is not None:
#             return document_content.text
#         else:
#             print("Error: Could not find document_content in the XML file.")
#             return None
#     except ET.ParseError as e:
#         print(f"Error parsing XML file: {e}")
#         return None
#     except FileNotFoundError:
#         print(f"Error: File not found: {file_path}")
#         return None
def extract_and_concatenate(input_string):
    # Extract the part that starts with 'M' followed by digits
    m_part_match = re.search(r'M\d+', input_string)
    
    # Extract the numbers after the underscore
    underscore_part_match = re.search(r'_(\d+)', input_string)
    
    if m_part_match and underscore_part_match:
        m_part = m_part_match.group()
        underscore_part = underscore_part_match.group(1)
        return m_part + underscore_part
    else:
        return "Invalid input format"


def extract_data(content):
    # Extract L and CM value
    l_pattern = r'L=(\d+M\s+\d+\.\d+CM)'
    l_match = re.search(l_pattern, content)
    l_value = l_match.group(1) if l_match else "Not found"

    # Extract U value
    u_pattern = r'U=(\d+\.\d+%)'
    u_match = re.search(u_pattern, content)
    u_value = u_match.group(1) if u_match else "Not found"

    # Extract PERIM value
    perim_pattern = r'PERIM=(\d+\.\d+CM)'
    perim_match = re.search(perim_pattern, content)
    perim_value = perim_match.group(1) if perim_match else "Not found"
    perim_value_wihtout_CM = perim_value.replace('CM','')

    # Extract LBMK value
    lbmk_pattern = r'LBMK:([\w-]+)'
    lbmk_match = re.search(lbmk_pattern, content)
    lbmk_value = lbmk_match.group(1) if lbmk_match else "Not found"

    length_M = l_value.split(' ')[0].replace('M', '')
    length_CM = l_value.split(' ')[
        1].replace('CM', '')
    length = (int(length_M)) + (float(length_CM)/100)
    lbmk = extract_and_concatenate(lbmk_value)

    return {
        "L": length,
        "U": u_value,
        "PERIM": perim_value_wihtout_CM,
        "LBMK": lbmk,
        "LBMK_full" : lbmk_value
    }

# Main function to tie it all together


def append_to_excel(data):
    # Convert the data dictionary to a pandas DataFrame
    new_df = pd.DataFrame([data])
    excel_path = r'data.xlsx'
    if os.path.exists(excel_path):
        # If file exists, read it and append new data
        existing_df = pd.read_excel(excel_path)
        updated_df = pd.concat([existing_df, new_df], ignore_index=True)
    else:
        # If file doesn't exist, create new DataFrame
        updated_df = new_df

    # Write the updated DataFrame to Excel
    updated_df.to_excel(excel_path, index=False)
    print(f"Data appended to {excel_path}")


def change_extension(file_path, new_extension):
    base_name, _ = os.path.splitext(file_path)
    new_file_path = base_name + "." + new_extension
    os.rename(file_path, new_file_path)
    return new_file_path
def find_and_edit_excel(data,appended_items,ws):
    # Define the column to search and the substring
    header_row = 1
# Populate headers-to-columns.
    fields = {}
    for cnum in range(1, ws.max_column + 1):
        field = ws.cell(row=header_row, column=cnum).value
        fields[field] = cnum
    substring = data['LBMK']  # Substring to search for
    percentage_style = NamedStyle(name='percentage_style', number_format='0.00%')
    number_style = NamedStyle(name='number_style', number_format='0.00')
    pattern = re.compile(r'^LN')
    linning = bool(pattern.match(data['LBMK_full']))
    # Iterate through the cells in the column
    for row_num in range(header_row + 1, ws.max_row + 1):
        job = ws.cell(row=row_num, column=fields['Job #']).value
        cut = ws.cell(row=row_num, column=fields['Cut #']).value
        concatenate = ws.cell(row=row_num, column=fields['Concatenate']).value
        job_cut = str(job) + str(cut)
        
        unique = job_cut + str(data['L']) +str(data['U'])
        
        # if (linning ==True) and ("Lining" in concatenate):
        #     pass
        # else:
        #     continue
        if (substring == job_cut) and (row_num not in appended_items):  # Check if the substring is in the cell's value
            # Edit cells in the same row
            if (linning ==True):
                if ("Lining" in concatenate):
                    pass
                else:
                    continue
            ws.cell(row=row_num, column=fields['Marker Length']).value = data['L']  # Modify cell B in the found row
            ws.cell(row=row_num, column=fields['Marker Length']).style= number_style
            ws.cell(row=row_num, column=fields['Marker Utilization']).value= data['U']
            ws.cell(row=row_num, column=fields['Marker Utilization']).style  = percentage_style
            ws.cell(row=row_num, column=fields['PARAMETER']).value= data['PERIM']
            ws.cell(row=row_num, column=fields['PARAMETER']).style  = number_style # Modify cell C in the found row
            appended_items.append(row_num)
            break  # Exit loop once the substring is found (remove if you want to find multiple occurrences)
    return row_num

def main(directory_path,excel_file_path):
    # Load the workbook
    wb = load_workbook(excel_file_path)
    data_json = {}
    appended_items = []
    # Select the active sheet
    sheet = wb['Sheet1']

    for index,filename in enumerate(os.listdir(directory_path)):
        print(f'=>Processing {index}. {filename}')
        file_path = os.path.join(directory_path, filename)
        if os.path.isfile(file_path):

            file_path = change_extension(file_path, 'txt')
            # content = read_file(file_path)
            with open(r'{}'.format(file_path), 'r') as file:
                content = file.read()
            if content:
                extracted_data = extract_data(content)
                data_json[f'{filename}'] = extracted_data
                for key, value in extracted_data.items():
                    print(f"{key}: {value}")
                insert = find_and_edit_excel(extracted_data,appended_items,sheet)
                extracted_data['unique_id'] = insert

            else:
                print("Failed to extract data due to file reading error.")
    with open('log.json','w') as json_logs:
        json_logs.write(json.dumps(data_json,indent=4))
    wb.save(f'updated_{os.path.basename(excel_file_path)}')
# Replace 'path/to/your/file.xml' with the actual path to your XML file

